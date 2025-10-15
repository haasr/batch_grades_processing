#!/usr/bin/env python3
"""
D2L Grades Management System - Command Line Interface

A unified CLI for querying, analyzing, and exporting student grades.
Combines Phase II (Query) and Phase III (Export/Report) functionality.
"""

import sys
from typing import Optional, Callable, List, Tuple
from queries import StudentQueries, SectionQueries, CohortQueries, GradeFormatter
from database.models import get_current_semester
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    # Suppress openpyxl warnings about missing thumbnails
    import warnings
    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
except ImportError:
    print("\n✗ openpyxl not installed!")
    print("Install with: pip install openpyxl --break-system-packages")

# Default semester - can be overridden
CURRENT_SEMESTER = get_current_semester()


class MenuItem:
    """Represents a single menu item."""
    
    def __init__(self, key: str, label: str, action: Callable, description: str = ""):
        self.key = key
        self.label = label
        self.action = action
        self.description = description
    
    def display(self) -> str:
        """Return formatted menu item for display."""
        return f"  {self.key}. {self.label}"


class MenuSystem:
    """Handles menu display and navigation."""
    
    def __init__(self, title: str, semester: str = CURRENT_SEMESTER):
        self.title = title
        self.semester = semester
        self.items: List[MenuItem] = []
        self.running = True

        year = semester[:4]
        semester_id = self.semester[-2:]
        if semester_id == '80':
            self.human_readable_semester = f"Fall {year}"
        elif semester_id == '10':
            self.human_readable_semester = f"Spring {year}"
        else:
            self.human_readable_semester = f"Summer {year}"

    def add_item(self, key: str, label: str, action: Callable, description: str = ""):
        """Add a menu item."""
        self.items.append(MenuItem(key, label, action, description))
    
    def add_separator(self):
        """Add a visual separator."""
        self.items.append(MenuItem("", "", lambda: None))
    
    def display(self):
        """Display the menu."""
        print("\n" + "="*80)
        print(self.title)
        print(f"Semester: {self.human_readable_semester}")
        print("="*80)
        print()
        
        for item in self.items:
            if item.key:  # Skip separators in display
                print(item.display())
            else:
                print()  # Empty line for separator
        
        print("\n  0. Exit")
        print("="*80)
    
    def get_choice(self) -> str:
        """Get user's menu choice."""
        while True:
            choice = input("\nEnter your choice: ").strip()
            if choice == "0":
                return "0"
            
            # Check if valid choice
            if any(item.key == choice for item in self.items if item.key):
                return choice
            
            print("✗ Invalid choice. Please try again.")
    
    def run(self):
        """Run the menu loop."""
        while self.running:
            self.display()
            choice = self.get_choice()
            
            if choice == "0":
                self.running = False
                print("\nGoodbye!")
                break
            
            # Execute the action
            for item in self.items:
                if item.key == choice:
                    print("\n" + "="*80)
                    try:
                        item.action()
                    except KeyboardInterrupt:
                        print("\n\n⚠️  Action cancelled by user.")
                    except Exception as e:
                        print(f"\n✗ Error: {str(e)}")
                        import traceback
                        traceback.print_exc()
                    print("="*80)
                    input("\n[Press Enter to continue]")
                    break


class CLIActions:
    """All CLI actions organized by category."""
    
    def __init__(self, semester: str = CURRENT_SEMESTER):
        self.semester = semester
    
    # =========================================================================
    # PHASE II: QUERY & VIEW ACTIONS
    # =========================================================================
    
    def lookup_student(self):
        """Look up a single student by username, org ID, or name."""
        print("STUDENT LOOKUP")
        print("-" * 80)
        print("\nSearch by:")
        print("  1. Username")
        print("  2. Org Defined ID")
        print("  3. Name (fuzzy search)")
        
        choice = input("\nEnter choice (1-3): ").strip()
        
        if choice == "1":
            username = input("Enter username: ").strip()
            student = StudentQueries.get_student_by_username(username, self.semester)
            if student:
                print(GradeFormatter.format_single_student(student))
                self._offer_export([student], f"student_{username}")
            else:
                print(f"\n✗ No student found with username '{username}'")
        
        elif choice == "2":
            org_id = input("Enter Org Defined ID: ").strip()
            student = StudentQueries.get_student_by_org_id(org_id, self.semester)
            if student:
                print(GradeFormatter.format_single_student(student))
                self._offer_export([student], f"student_{org_id}")
            else:
                print(f"\n✗ No student found with Org ID '{org_id}'")
        
        elif choice == "3":
            name = input("Enter name (first, last, or both): ").strip()
            students = StudentQueries.search_students_by_name(name, self.semester, limit=20)
            if students:
                print(f"\n✓ Found {len(students)} student(s) matching '{name}':")
                print(GradeFormatter.format_student_list(students))
                self._offer_export(students, f"search_{name.replace(' ', '_')}")
            else:
                print(f"\n✗ No students found matching '{name}'")
        
        else:
            print("✗ Invalid choice")
    
    def view_section(self):
        """View all students in a specific section."""
        print("SECTION VIEW")
        print("-" * 80)
        
        # List available sections
        sections = SectionQueries.list_available_sections(self.semester)
        if not sections:
            print("\n✗ No sections found in database")
            return
        
        # Group by type
        lab_sections = [s for s in sections if s['course_type'] == 'LAB']
        lecture_sections = [s for s in sections if s['course_type'] == 'LECTURE']
        
        print("\nLab sections:")
        for idx, s in enumerate(lab_sections, 1):
            print(f"  {idx}. {s['course_name']}-{s['section']}")
        
        print(f"\nLecture sections:")
        # Continue numbering from where lab sections left off
        start_num = len(lab_sections) + 1
        for idx, s in enumerate(lecture_sections, start_num):
            print(f"  {idx}. {s['course_name']}-{s['section']}")
        
        # Get choice
        try:
            choice = int(input("\nEnter section number (or 0 to cancel): ").strip())
            if choice == 0:
                return
            
            if 1 <= choice <= len(sections):
                section = sections[choice + 1]
                students = SectionQueries.get_section_grades(
                    section['course_name'],
                    section['section'],
                    self.semester
                )
                
                if students:
                    section_name = f"{section['course_name']}-{section['section']}"
                    print(GradeFormatter.format_section_summary(students, section_name))
                    self._offer_export(students, f"section_{section_name}")
                else:
                    print(f"\n✗ No students found in section")
            else:
                print("✗ Invalid section number")
        
        except ValueError:
            print("✗ Invalid input")
    
    def compare_cohorts(self):
        """Compare in-person vs online cohorts."""
        print("COHORT COMPARISON")
        print("-" * 80)
        
        # Get statistics for both
        inperson_stats = CohortQueries.get_cohort_statistics(self.semester, 'inperson')
        online_stats = CohortQueries.get_cohort_statistics(self.semester, 'online')
        
        # Display both
        print(GradeFormatter.format_cohort_statistics(inperson_stats, 'inperson'))
        print("\n")
        print(GradeFormatter.format_cohort_statistics(online_stats, 'online'))
        
        # Side-by-side comparison
        print("\n" + "="*80)
        print("SIDE-BY-SIDE COMPARISON")
        print("="*80)
        print(f"{'Metric':<35} {'In-Person':>15} {'Online':>15} {'Difference':>15}")
        print("-"*80)
        
        metrics = [
            ('Total Students', 'total_students', 'd'),
            ('Students with DCA', 'students_with_dca', 'd'),
            ('Avg Lab Grade', 'avg_lab', '.2f'),
            ('Avg Quizzes Grade', 'avg_quizzes', '.2f'),
            ('Avg Exit Tickets Grade', 'avg_exit_tickets', '.2f'),
            ('Avg Pre-Final Grade', 'avg_overall_pre', '.2f'),
            ('Avg Post-Final Grade', 'avg_overall_post', '.2f'),
            ('Passing Rate (Pre)', 'passing_rate_pre', '.1f'),
            ('Passing Rate (Post)', 'passing_rate_post', '.1f'),
        ]
        
        for label, key, fmt in metrics:
            ip_val = inperson_stats[key]
            ol_val = online_stats[key]
            diff = ip_val - ol_val
            
            if fmt == 'd':
                print(f"{label:<35} {ip_val:>15} {ol_val:>15} {diff:>+15}")
            else:
                unit = '%' if 'rate' in key.lower() or 'avg' in key.lower() else ''
                print(f"{label:<35} {ip_val:>14{fmt}}{unit} {ol_val:>14{fmt}}{unit} {diff:>+14{fmt}}{unit}")
        
        print("="*80)
        
        # Offer to export both cohorts
        print("\n📊 Export Options:")
        print("  1. Export in-person students")
        print("  2. Export online students")
        print("  3. Export both")
        print("  0. Don't export")
        
        choice = input("\nEnter choice: ").strip()
        
        if choice == "1":
            students = CohortQueries.get_inperson_students(self.semester)
            self._export_to_csv(students, f"inperson_{self.semester}")
        elif choice == "2":
            students = CohortQueries.get_online_students(self.semester)
            self._export_to_csv(students, f"online_{self.semester}")
        elif choice == "3":
            inperson = CohortQueries.get_inperson_students(self.semester)
            online = CohortQueries.get_online_students(self.semester)
            self._export_to_csv(inperson, f"inperson_{self.semester}")
            self._export_to_csv(online, f"online_{self.semester}")
    
    def find_at_risk(self):
        """Find students who are at risk based on their grades."""
        print("AT-RISK STUDENTS")
        print("-" * 80)
        
        print("\nSelect mode:")
        print("  1. Pre-final (students with pre-final < 60%)")
        print("  2. Post-final (students passing pre-final but failing post-final)")

        choice = input("\nEnter choice (1-2): ").strip()

        all_students = CohortQueries.get_all_students(self.semester)
        
        if choice == "1":
            # Before final project - show students currently failing
            print("\nFinding students with pre-final grade < 60%...")
            at_risk = [
                s for s in all_students
                if s['overall_pre_final'] < 60.0
            ]

            if at_risk:
                print(f"\n⚠️  Found {len(at_risk)} at-risk student(s) (pre-final < 60%):")
                print(GradeFormatter.format_student_list(at_risk, show_full_grades=True))

                # Show breakdown by grade range
                print("\n📊 Grade Distribution:")
                ranges = [
                    (0, 50, "Critically at-risk"),
                    (50, 55, "Very at-risk"),
                    (55, 60, "At-risk")
                ]
                for low, high, label in ranges:
                    count = len([s for s in at_risk if low <= s['overall_pre_final'] < high])
                    if count > 0:
                        print(f"  {label} ({low}-{high}%): {count} students")

                self._offer_export(at_risk, f"at_risk_pre_final_{self.semester}")
            else:
                print("\n✓ No at-risk students found! Everyone is passing.")

        elif choice == "2":
            # After final project - show students who were passing but now failing
            print("\nFinding students with pre-final ≥ 60% but post-final < 60%...")
            at_risk = [
                s for s in all_students
                if s['overall_pre_final'] >= 60.0 and s['overall_post_final'] < 60.0
            ]

            if at_risk:
                print(f"\n⚠️  Found {len(at_risk)} at-risk student(s):")
                print("(These students were passing before the final but are now failing)")
                print(GradeFormatter.format_student_list(at_risk, show_full_grades=True))

                # Show who hasn't submitted DCA
                no_dca = [s for s in at_risk if not s['has_final_project']]
                if no_dca:
                    print(f"\n⚠️  {len(no_dca)} of these students have NOT submitted their DCA!")
                    print("   (They may still be able to pass if they submit)")

                self._offer_export(at_risk, f"at_risk_post_final_{self.semester}")
            else:
                print("\n✓ No students lost their passing status after the final!")

        else:
            print("✗ Invalid choice")
    
    def find_missing_dca(self):
        """Find students who haven't submitted the DCA."""
        print("MISSING DCA REPORT")
        print("-" * 80)
        
        all_students = CohortQueries.get_all_students(self.semester)
        missing_dca = [s for s in all_students if not s['has_final_project']]
        
        if missing_dca:
            print(f"\n⚠️  Found {len(missing_dca)} student(s) without DCA:")
            print(GradeFormatter.format_student_list(missing_dca))
            self._offer_export(missing_dca, f"missing_dca_{self.semester}")
        else:
            print("\n✓ All students have submitted their DCA!")
    
    # =========================================================================
    # PHASE III: EXPORT & REPORT ACTIONS (Stubs for now)
    # =========================================================================
    
    def export_lab_workbook(self):
        """Generate giant Excel workbook for lab instructors."""
        print("LAB INSTRUCTOR WORKBOOK")
        print("-" * 80)

        print("\n📊 Creating comprehensive workbook...")
        print("   This may take a minute...\n")

        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # =====================================================================
        # PART 1: LAB SECTION SHEETS
        # =====================================================================
        print("📁 Creating lab section sheets...")

        lab_sections = SectionQueries.list_available_sections(self.semester, course_type='LAB')

        for idx, section in enumerate(lab_sections, 1):
            section_name = f"{section['course_name']}-{section['section']}"

            # Get students for this section
            students = SectionQueries.get_lab_section_grades(
                section['course_name'],
                section['section'],
                self.semester
            )

            if not students:
                continue

            # Create sheet (Excel limits sheet names to 31 chars)
            sheet_name = section_name[:31]
            ws = wb.create_sheet(sheet_name)

            # Add title row
            ws.merge_cells('A1:F1')
            title_cell = ws['A1']
            title_cell.value = f"Lab Section: {section_name}"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center')

            # Add headers
            headers = ['Last Name', 'First Name', 'Username', 'Email', 'Lab Average', 'DCA Score']
            ws.append([])  # Empty row
            ws.append(headers)

            # Style headers
            for cell in ws[3]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            # Add student data
            for student in sorted(students, key=lambda s: s['last_name']):
                ws.append([
                    student['last_name'],
                    student['first_name'],
                    student['username'],
                    student['email'],
                    f"{student['lab_average']:.2f}",
                    f"{student['dca_score']:.2f}" if student['has_final_project'] else "Not Submitted",
                ])

            # Auto-size columns
            for col_idx, column in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Freeze header row
            ws.freeze_panes = 'A4'

            print(f"  ✓ {section_name} ({len(students)} students)")

        # =====================================================================
        # PART 2: LECTURE COHORT SHEETS
        # =====================================================================
        print("\n📊 Creating lecture cohort sheets...")

        # In-Person Cohort
        inperson_students = CohortQueries.get_inperson_students(self.semester)
        self._create_lecture_sheet(wb, "In-Person Lecture", inperson_students,
                                   header_fill, header_font, border)
        print(f"  ✓ In-Person Lecture ({len(inperson_students)} students)")

        # Online Cohort
        online_students = CohortQueries.get_online_students(self.semester)
        self._create_lecture_sheet(wb, "Online Lecture", online_students,
                                   header_fill, header_font, border)
        print(f"  ✓ Online Lecture ({len(online_students)} students)")

        # =====================================================================
        # PART 3: SUMMARY STATISTICS SHEETS
        # =====================================================================
        print("\n📈 Creating summary statistics sheets...")

        inperson_stats = CohortQueries.get_cohort_statistics(self.semester, 'inperson')
        self._create_summary_sheet(wb, "In-Person Summary", inperson_stats,
                                   header_fill, header_font)
        print(f"  ✓ In-Person Summary")


        online_stats = CohortQueries.get_cohort_statistics(self.semester, 'online')
        self._create_summary_sheet(wb, "Online Summary", online_stats,
                                   header_fill, header_font)
        print(f"  ✓ Online Summary")

        # =====================================================================
        # SAVE WORKBOOK
        # =====================================================================
        filename = f'lab_workbook_{self.semester}.xlsx'
        wb.save(filename)

        print(f"\n✓ Created {filename}")
        print(f"📊 Total sheets: {len(wb.sheetnames)}")
        print(f"   • {len(lab_sections)} lab section sheets")
        print(f"   • 2 lecture cohort sheets")
        print(f"   • 2 summary statistics sheets")
        print(f"\n💡 This workbook is ready to distribute to lab instructors!")
    
    def _create_lecture_sheet(self, wb, sheet_name, students, header_fill, header_font, border):
        """Create a lecture cohort sheet with full grade details."""
        ws = wb.create_sheet(sheet_name)

        # Add title
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = sheet_name
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        # Add headers
        headers = [
            'Last Name', 'First Name', 'Username', 'Email',
            'Lab Avg', 'Quiz Avg', 'Exit Ticket Avg', 'DCA',
            'Pre-Final', 'Post-Final', 'Lab Section'
        ]
        ws.append([])  # Empty row
        ws.append(headers)

        # Style headers
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Add student data
        for student in sorted(students, key=lambda s: s['last_name']):
            ws.append([
                student['last_name'],
                student['first_name'],
                student['username'],
                student['email'],
                f"{student['lab_average']:.2f}",
                f"{student['quizzes_average']:.2f}",
                f"{student['exit_tickets_average']:.2f}",
                f"{student['dca_score']:.2f}" if student['has_final_project'] else "Not Submitted",
                f"{student['overall_pre_final']:.2f}",
                f"{student['overall_post_final']:.2f}",
                student['lab_section'] or 'N/A',
            ])

        # Auto-size columns
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A4'

    def _create_summary_sheet(self, wb, sheet_name, stats, header_fill, header_font):
        """Create a summary statistics sheet with grade distribution."""
        ws = wb.create_sheet(sheet_name)

        # Add title
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.value = sheet_name
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        ws.append([])  # Empty row

        # Add statistics
        stats_data = [
            ['Metric', 'Value'],
            ['Total Students', stats['total_students']],
            ['Students with DCA', stats['students_with_dca']],
            ['Students with Lab Grade', stats['students_with_lab']],
            ['Students with Quiz Grade', stats['students_with_quizzes']],
            ['Students with Exit Ticket Grade', stats['students_with_exit_tickets']],
            [],
            ['Average Lab Grade', f"{stats['avg_lab']:.2f}%"],
            ['Average Quiz Grade', f"{stats['avg_quizzes']:.2f}%"],
            ['Average Exit Ticket Grade', f"{stats['avg_exit_tickets']:.2f}%"],
            ['Average DCA Score', f"{stats['avg_dca']:.2f}%"],
            [],
            ['Average Pre-Final Grade', f"{stats['avg_overall_pre']:.2f}%"],
            ['Average Post-Final Grade', f"{stats['avg_overall_post']:.2f}%"],
            [],
            ['Passing Rate (Pre-Final)', f"{stats['passing_rate_pre']:.1f}%"],
            ['Passing Rate (Post-Final)', f"{stats['passing_rate_post']:.1f}%"],
        ]

        for row in stats_data:
            ws.append(row)

        # Style header row
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Auto-size columns
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

        # Add grade distribution for post-final grades
        ws.append([])
        ws.append(['Grade Distribution (Post-Final)', 'Count'])
        dist_header_row = ws.max_row
        ws['A' + str(dist_header_row)].font = Font(bold=True, size=12)
        ws['A' + str(dist_header_row)].fill = header_fill
        ws['B' + str(dist_header_row)].fill = header_fill
        ws['A' + str(dist_header_row)].font = Font(bold=True, color="FFFFFF")
        ws['B' + str(dist_header_row)].font = Font(bold=True, color="FFFFFF")

        # Get actual distribution counts from grade_distribution in stats
        if 'grade_distribution' in stats:
            dist = stats['grade_distribution']
            ws.append(['A (90-100)', dist.get('A', 0)])
            ws.append(['B (80-89)', dist.get('B', 0)])
            ws.append(['C (70-79)', dist.get('C', 0)])
            ws.append(['D (60-69)', dist.get('D', 0)])
            ws.append(['F (<60)', dist.get('F', 0)])
        else:
            # Fallback if distribution not available
            ws.append(['A (90-100)', 'N/A'])
            ws.append(['B (80-89)', 'N/A'])
            ws.append(['C (70-79)', 'N/A'])
            ws.append(['D (60-69)', 'N/A'])
            ws.append(['F (<60)', 'N/A'])

    def export_d2l_espr(self):
        """Export ESPR grades in D2L import format."""
        print("D2L ESPR GRADES EXPORT")
        print("-" * 80)
        print("\nESPR = End of Semester Performance Report (Midterm Grades)")
        print("This exports the pre-final grades for D2L import.")

        print("\nSelect cohort:")
        print("  1. In-person students only")
        print("  2. Online students only")
        print("  3. All students")
        print("  0. Cancel")

        choice = input("\nEnter choice (1-3): ").strip()

        if choice == "0":
            return
        elif choice == "1":
            students = CohortQueries.get_inperson_students(self.semester)
            cohort_label = "inperson"
        elif choice == "2":
            students = CohortQueries.get_online_students(self.semester)
            cohort_label = "online"
        elif choice == "3":
            students = CohortQueries.get_all_students(self.semester)
            cohort_label = "all"
        else:
            print("✗ Invalid choice")
            return

        if not students:
            print("\n✗ No students found")
            return

        # D2L ESPR import format
        # Expected columns: Username, OrgDefinedId, ESPR Grade
        filename = f'd2l_espr_{cohort_label}_{self.semester}.csv'

        import csv
        try:
            with open(filename, 'w', newline='') as f:
                writer = csv.writer(f)

                # D2L header row (with End-of-Line Indicator)
                writer.writerow(['Username', 'OrgDefinedId', 'ESPR Grade', 'End-of-Line Indicator'])

                for student in students:
                    # Use pre-final grade for ESPR
                    espr_grade = student['overall_pre_final']

                    writer.writerow([
                        student['username'],
                        student['org_defined_id'],
                        f"{espr_grade:.2f}",
                        '#'  # D2L requires this for parsing
                    ])

            print(f"\n✓ Exported {len(students)} ESPR grades to {filename}")
            print(f"📤 Ready to import into D2L!")
            print(f"\n💡 Import instructions:")
            print(f"   1. Go to D2L course site")
            print(f"   2. Grades → Import")
            print(f"   3. Select '{filename}'")
            print(f"   4. Map to ESPR Grade column")

        except Exception as e:
            print(f"\n✗ Export failed: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def export_d2l_final(self):
        """Export final grades in D2L import format."""
        print("D2L FINAL GRADES EXPORT")
        print("-" * 80)
        print("\nThis exports the post-final grades for D2L import.")

        print("\nSelect cohort:")
        print("  1. In-person students only")
        print("  2. Online students only")
        print("  3. All students")
        print("  0. Cancel")

        choice = input("\nEnter choice (1-3): ").strip()

        if choice == "0":
            return
        elif choice == "1":
            students = CohortQueries.get_inperson_students(self.semester)
            cohort_label = "inperson"
        elif choice == "2":
            students = CohortQueries.get_online_students(self.semester)
            cohort_label = "online"
        elif choice == "3":
            students = CohortQueries.get_all_students(self.semester)
            cohort_label = "all"
        else:
            print("✗ Invalid choice")
            return

        if not students:
            print("\n✗ No students found")
            return

        # D2L final grades import format
        # Expected columns: Username, OrgDefinedId, End-of-Term Grade
        filename = f'd2l_final_{cohort_label}_{self.semester}.csv'

        import csv
        try:
            with open(filename, 'w', newline='') as f:
                writer = csv.writer(f)

                # D2L header row (with End-of-Line Indicator)
                writer.writerow(['Username', 'OrgDefinedId', 'End-of-Term Grade', 'End-of-Line Indicator'])

                for student in students:
                    # Use post-final grade for final
                    final_grade = student['overall_post_final']

                    writer.writerow([
                        student['username'],
                        student['org_defined_id'],
                        f"{final_grade:.2f}",
                        '#'  # D2L requires this for parsing
                    ])

            print(f"\n✓ Exported {len(students)} final grades to {filename}")
            print(f"📤 Ready to import into D2L!")

            # Show grade distribution summary
            passing = len([s for s in students if s['overall_post_final'] >= 60.0])
            failing = len(students) - passing

            print(f"\n📊 Grade Summary:")
            print(f"   Passing (≥60%): {passing} students ({passing/len(students)*100:.1f}%)")
            print(f"   Failing (<60%): {failing} students ({failing/len(students)*100:.1f}%)")

            print(f"\n💡 Import instructions:")
            print(f"   1. Go to D2L course site")
            print(f"   2. Grades → Import")
            print(f"   3. Select '{filename}'")
            print(f"   4. Map to End-of-Term Grade column")

        except Exception as e:
            print(f"\n✗ Export failed: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def custom_export(self):
        """Custom CSV export with user-selected filters."""
        print("CUSTOM EXPORT")
        print("-" * 80)
        print("\n🚧 Coming soon in Phase III!")
        print("\nThis will allow you to:")
        print("  • Select specific sections or cohorts")
        print("  • Choose which grade components to include")
        print("  • Filter by grade thresholds")
        print("  • Export to CSV with custom formatting")
    
    def grade_distribution(self):
        """Show grade distribution statistics."""
        print("GRADE DISTRIBUTION")
        print("-" * 80)
        print("\n🚧 Coming soon in Phase III!")
        print("\nThis will show:")
        print("  • Histogram of grade distributions")
        print("  • Percentile breakdowns")
        print("  • Comparison across sections")
    
    # =========================================================================
    # HELPER METHODS
    # =========================================================================
    
    def _offer_export(self, students: List, default_filename: str):
        """Offer to export results to CSV."""
        export = input("\n📥 Export to CSV? (y/n): ").strip().lower()
        if export == 'y':
            self._export_to_csv(students, default_filename)
    
    def _export_to_csv(self, students: List, filename: str):
        """Export students to CSV file."""
        import csv
        
        if not students:
            print("✗ No data to export")
            return
        
        # Ensure .csv extension
        if not filename.endswith('.csv'):
            filename += '.csv'
        
        try:
            with open(filename, 'w', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=students[0].keys())
                writer.writeheader()
                writer.writerows(students)
            
            print(f"\n✓ Exported {len(students)} student(s) to {filename}")
        
        except Exception as e:
            print(f"\n✗ Export failed: {str(e)}")


def main():
    """Main CLI entry point."""
    # Parse command line args
    semester = CURRENT_SEMESTER
    if len(sys.argv) > 1:
        semester = sys.argv[1]
    
    # Create menu system
    menu = MenuSystem("CSCI 1100 Grades Processing System", semester)
    actions = CLIActions(semester)
    
    # Add Phase II items
    menu.add_item("1", "⌕ Lookup student", actions.lookup_student)
    menu.add_item("2", "§ View section grades", actions.view_section)
    menu.add_item("3", "- Compare cohorts (in-person vs online)", actions.compare_cohorts)
    menu.add_item("4", "⚑ Find at-risk students", actions.find_at_risk)
    menu.add_item("5", "⚐ Find students missing DCA", actions.find_missing_dca)
    
    menu.add_separator()

    # Add Phase III items (stubs for now)
    menu.add_item("6", "⎙ Generate lab instructor workbook", actions.export_lab_workbook)
    menu.add_item("7", "↥ Export ESPR grades (D2L format)", actions.export_d2l_espr)
    menu.add_item("8", "↥ Export final grades (D2L format)", actions.export_d2l_final)
    
    # Run the menu
    menu.run()


if __name__ == '__main__':
    main()
